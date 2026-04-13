from __future__ import annotations

"""Reader per file Excel.

Converte ogni worksheet in `RawSheet` mantenendo metadati grafici minimi
(grassetto, colore di sfondo, dimensione font, merge).
"""

from pathlib import Path

import openpyxl

from quoro.models import RawCell, RawSheet


def _cell_bg_color(cell) -> str | None:
    """Estrae il colore di sfondo della cella, se disponibile."""

    fill = cell.fill
    if fill and fill.fgColor and fill.fgColor.type != "none":
        rgb = fill.fgColor.rgb
        if rgb and rgb != "00000000":
            return rgb
    return None


def _cell_font_size(cell) -> float | None:
    """Estrae la dimensione font della cella, se disponibile."""

    if cell.font and cell.font.size:
        return float(cell.font.size)
    return None


def read_excel(path: Path) -> list[RawSheet]:
    """Legge un workbook Excel e restituisce un `RawSheet` per ogni worksheet.

    Le aree merge vengono espanse in lettura replicando il valore della cella
    top-left su tutte le celle del range, cosi l'Analyzer vede una griglia
    completa e non celle vuote artificiali.
    """

    wb = openpyxl.load_workbook(path, data_only=True)
    sheets: list[RawSheet] = []

    for ws in wb.worksheets:
        # Build merged cell lookup: maps (row, col) -> value of top-left cell
        merged_cells: dict[tuple[int, int], str] = {}
        for merged_range in ws.merged_cells.ranges:
            top_left = ws.cell(merged_range.min_row, merged_range.min_col)
            val = str(top_left.value) if top_left.value is not None else ""
            for row in range(merged_range.min_row, merged_range.max_row + 1):
                for col in range(merged_range.min_col, merged_range.max_col + 1):
                    if not (
                        row == merged_range.min_row and col == merged_range.min_col
                    ):
                        merged_cells[(row, col)] = val

        rows: list[list[RawCell]] = []
        for row in ws.iter_rows():
            cells: list[RawCell] = []
            for cell in row:
                row_num = cell.row or 0
                col_num = cell.column or 0
                coord: tuple[int, int] = (row_num, col_num)
                is_merged = coord in merged_cells
                if is_merged:
                    value = merged_cells[coord]
                else:
                    value = str(cell.value) if cell.value is not None else ""

                bold = bool(cell.font and cell.font.bold)
                bg_color = _cell_bg_color(cell)
                font_size = _cell_font_size(cell)

                cells.append(
                    RawCell(
                        value=value.strip(),
                        bold=bold,
                        bg_color=bg_color,
                        font_size=font_size,
                        merged=is_merged,
                    )
                )
            rows.append(cells)

        sheets.append(RawSheet(name=ws.title, rows=rows, separator=None))

    return sheets
